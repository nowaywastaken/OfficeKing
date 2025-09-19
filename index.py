#!/usr/bin/env python3
"""一键运行：Office/PDF 转文本并检索命中的学号/姓名，导出 Excel。

用法：
  python index.py

行为：
  - 启动即从根目录 `config.yml` 读取全部参数
  - 使用 MarkItDown 转换 Word/Excel/PPT（含老旧格式）为 Markdown 文本
  - 对 PDF 同时进行：嵌入文本提取 + Tesseract OCR，再与 MarkItDown 转换结果合并
  - 只搜索配置中给定的学生映射（姓名→学号）
  - 若命中，记录：命中的内容、文件名、保存路径；导出到 Output/时间戳(含提示语) 目录
  - 同时将命中的源文件复制到该目录
  - 全程写入同一个 `log.txt` 并同步输出到终端
"""

from __future__ import annotations

import logging
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple, Set
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd  # type: ignore[import-not-found]

from activity_scanner.config import (
    CONFIG_ROOT,
    LOG_LEVEL,
    STUDENT_ID_MAP,
    LOG_SUPPRESS_LOGGER_PREFIXES,
    LOG_SUPPRESS_MESSAGE_CONTAINS,
    LOG_CLEANUP_REMOVE_WARNING_LINES,
    PDF_WORKERS,
)
from activity_scanner.extractors import SUPPORTED_EXTENSIONS, read_text_from_path


def setup_logging(log_path: Path, level_name: str) -> None:
    """Configure logging to console and a single append-only file.

    Adds suppression filters for known noisy third-party warnings controlled by
    config.yml to ensure the log file stays clean without reducing functionality.
    """

    level = getattr(logging, str(level_name).upper(), logging.INFO)
    logger = logging.getLogger()
    logger.setLevel(level)
    logger.handlers.clear()

    # Build a suppression filter based on config values
    class _SuppressFilter(logging.Filter):
        def filter(self, record: logging.LogRecord) -> bool:  # type: ignore[override]
            try:
                # Drop by logger name prefixes
                for prefix in LOG_SUPPRESS_LOGGER_PREFIXES:
                    if prefix and record.name.startswith(prefix):
                        return False
                # Drop by substring match in message
                msg = record.getMessage()
                for sub in LOG_SUPPRESS_MESSAGE_CONTAINS:
                    if sub and sub in msg:
                        return False
            except Exception:
                # Never fail logging due to filtering errors
                pass
            return True

    suppress_filter = _SuppressFilter()

    fmt = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # Optional cleanup: remove historical WARNING lines before opening file handler
    try:
        if LOG_CLEANUP_REMOVE_WARNING_LINES and log_path.exists():
            orig = log_path.read_text(encoding="utf-8", errors="ignore").splitlines()
            kept = [ln for ln in orig if "[WARNING]" not in ln]
            if len(kept) != len(orig):
                tmp = "\n".join(kept) + ("\n" if kept else "")
                log_path.write_text(tmp, encoding="utf-8")
    except Exception:
        # Do not fail if cleanup cannot complete
        pass

    ch = logging.StreamHandler()
    ch.setLevel(level)
    ch.setFormatter(fmt)
    ch.addFilter(suppress_filter)
    logger.addHandler(ch)

    fh = logging.FileHandler(log_path, mode="a", encoding="utf-8")
    fh.setLevel(level)
    fh.setFormatter(fmt)
    fh.addFilter(suppress_filter)
    logger.addHandler(fh)

    # Additionally, raise levels for specified noisy loggers
    for prefix in LOG_SUPPRESS_LOGGER_PREFIXES:
        try:
            if prefix:
                logging.getLogger(prefix).setLevel(max(level, logging.ERROR))
        except Exception:
            pass


def _coerce_paths(values: Iterable[str]) -> List[Path]:
    """Return existing files/dirs as Path list (deduplicated, absolute)."""

    out: List[Path] = []
    seen: set[str] = set()
    for raw in values:
        p = (CONFIG_ROOT / str(raw)).resolve() if not os.path.isabs(str(raw)) else Path(raw).resolve()
        if p.exists():
            key = str(p)
            if key not in seen:
                seen.add(key)
                out.append(p)
        else:
            logging.warning("[input] 路径不存在: %s", p)
    return out


def collect_supported_files(inputs: Iterable[Path], exts: Iterable[str]) -> List[Path]:
    """Recursively collect supported files under the given inputs."""

    supported = {e.lower() for e in exts}
    results: List[Path] = []
    for item in inputs:
        if item.is_file():
            if item.suffix.lower() in supported:
                results.append(item.resolve())
        elif item.is_dir():
            for root, _dirs, files in os.walk(item):
                for name in files:
                    if os.path.splitext(name)[1].lower() in supported:
                        results.append((Path(root) / name).resolve())
        else:
            logging.warning("[input] 非文件/目录: %s", item)
    # stable order + dedup
    as_str = sorted(set(str(p) for p in results))
    return [Path(s) for s in as_str]


def _choose_workers(config_workers: int) -> int:
    """Choose a small thread pool size honoring config value when > 0.

    The goal is "small-scale" concurrency. If config is <= 0, pick a conservative
    value based on CPU count, capped at 4.
    """

    try:
        n = int(config_workers)
    except Exception:
        n = 0
    if n > 0:
        return n
    cpu = os.cpu_count() or 2
    return min(4, max(2, cpu))


def _build_name_resolver(name_to_id: Dict[str, str]) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Return (id->name, alt->name) mappings for quick canonicalization.

    The "alt" mapping covers dotless variants of names that contain the middle dot '·'.
    """

    id_to_name: Dict[str, str] = {sid: name for name, sid in name_to_id.items()}
    alt_to_name: Dict[str, str] = {}
    for name in name_to_id.keys():
        if "·" in name:
            alt_to_name[name.replace("·", "")] = name
    return id_to_name, alt_to_name


def _canonicalize_token(token: str, name_to_id: Dict[str, str], id_to_name: Dict[str, str], alt_to_name: Dict[str, str]) -> str:
    """Map a hit token to a canonical student name when possible.

    - If token equals a student ID -> return the corresponding name
    - If token equals a name (exact) -> return that name
    - If token equals an alt variant (dotless) -> return the dotted canonical name
    Otherwise return the token unchanged.
    """

    if token in id_to_name:
        return id_to_name[token]
    if token in name_to_id:
        return token
    if token in alt_to_name:
        return alt_to_name[token]
    return token


def _process_one_file(path: Path, name_to_id: Dict[str, str], id_to_name: Dict[str, str], alt_to_name: Dict[str, str]) -> Tuple[Path, Set[str]]:
    """Read text and return the set of canonical names found in this file."""

    try:
        text = read_text_from_path(str(path))
    except ImportError as exc:
        # Bubble up ImportError so caller can terminate with a clear message
        raise exc
    except Exception as exc:
        logging.warning("读取失败 %s: %s", path, exc)
        return path, set()

    names: Set[str] = set()
    for token in _search_hits(text or "", name_to_id):
        canon = _canonicalize_token(token, name_to_id, id_to_name, alt_to_name)
        # Only keep tokens that resolve to known names; this also deduplicates
        if canon in name_to_id:
            names.add(canon)
    return path, names


def _autosize_excel_columns(xlsx_path: Path, sheet_name: str) -> None:
    """Auto-fit column widths in-place using openpyxl."""

    try:
        import openpyxl  # type: ignore[import-not-found]
        from openpyxl.utils import get_column_letter  # type: ignore[import-not-found]
        import unicodedata
    except Exception as exc:  # pragma: no cover - environment dependent
        logging.warning("自动列宽失败(依赖缺失): %s", exc)
        return

    def _disp_len(s: str) -> float:
        # Treat full-width/East Asian chars as width 2
        total = 0.0
        for ch in s:
            ea = unicodedata.east_asian_width(ch)
            total += 2.0 if ea in ("W", "F") else 1.0
        return total

    try:
        wb = openpyxl.load_workbook(filename=str(xlsx_path))
        ws = wb[sheet_name]
        # Compute max width per column including header
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
            header = col_cells[0].value
            max_len = _disp_len(str(header)) if header is not None else 0.0
            for cell in col_cells[1:]:
                val = cell.value
                if val is None:
                    continue
                max_len = max(max_len, _disp_len(str(val)))
            # Add padding; Excel uses roughly 1 unit per char
            width = min(100.0, max(8.0, max_len + 2.0))
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
        wb.save(str(xlsx_path))
    except Exception as exc:
        logging.warning("自动列宽设置失败: %s", exc)


def _load_config_values() -> Tuple[List[Path], Path, str, str]:
    """Load required values from config.yml.

    Returns: (inputs, output_root, output_folder_format, report_filename)
    """

    import yaml  # type: ignore[import-not-found]

    cfg_path = CONFIG_ROOT / "config.yml"
    data = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}

    # All parameters are sourced from config.yml, no in-code defaults.
    input_paths = data.get("input_paths")
    output_root = data.get("output_root_dir")
    folder_fmt = data.get("output_folder_format")
    report_name = data.get("report_filename")

    if not isinstance(input_paths, list) or not input_paths:
        raise ValueError("config.yml: input_paths 必须是非空列表")
    if not isinstance(output_root, str) or not output_root:
        raise ValueError("config.yml: output_root_dir 必须是非空字符串")
    if not isinstance(folder_fmt, str) or not folder_fmt:
        raise ValueError("config.yml: output_folder_format 必须是非空字符串")
    if not isinstance(report_name, str) or not report_name.endswith(".xlsx"):
        raise ValueError("config.yml: report_filename 必须是以 .xlsx 结尾的文件名")

    inputs = _coerce_paths(input_paths)
    if not inputs:
        raise FileNotFoundError("input_paths 指向的文件/目录均不存在")

    return inputs, (CONFIG_ROOT / output_root).resolve(), folder_fmt, report_name


def _ensure_output_folder(root: Path, folder_fmt: str) -> Path:
    """Create the timestamped output folder and return it."""

    # Allow strftime tokens in folder_fmt
    name = datetime.now().strftime(folder_fmt)
    target = (root / name).resolve()
    target.mkdir(parents=True, exist_ok=True)
    return target


def _search_hits(text: str, name_to_id: Dict[str, str]) -> List[str]:
    """Return a list of matched raw tokens (names and IDs)."""

    matches: List[str] = []
    for name, sid in name_to_id.items():
        if name and name in text:
            matches.append(name)
        if sid and sid in text:
            matches.append(sid)
        # name without middle dot variant
        if "·" in name:
            alt = name.replace("·", "")
            if alt and alt in text and alt not in matches:
                matches.append(alt)
    return matches


def main() -> int:
    """Run the full scan using YAML config only."""

    log_path = CONFIG_ROOT / "log.txt"
    setup_logging(log_path=log_path, level_name=LOG_LEVEL)

    logging.info("读取 config.yml 并初始化")
    try:
        inputs, out_root, folder_fmt, report_name = _load_config_values()
    except Exception as exc:
        logging.error("配置读取失败: %s", exc)
        return 1

    logging.info("输入路径: %s", [str(p) for p in inputs])
    files = collect_supported_files(inputs, SUPPORTED_EXTENSIONS)
    if not files:
        logging.info("未找到可处理的文件 (支持: %s)", sorted(SUPPORTED_EXTENSIONS))
        return 0

    # Resolve roster and helpers
    name_to_id = dict(STUDENT_ID_MAP)
    id_to_name, alt_to_name = _build_name_resolver(name_to_id)
    logging.info("学生名单条目数: %d", len(name_to_id))

    # Prepare output
    out_dir = _ensure_output_folder(out_root, folder_fmt)
    report_path = out_dir / report_name
    logging.info("输出目录: %s", out_dir)

    rows: List[Dict[str, str]] = []
    matched_files: List[Path] = []

    total = len(files)
    worker_count = _choose_workers(PDF_WORKERS)
    logging.info("开始并发处理文件: %d (workers=%d)", total, worker_count)

    # Submit tasks
    with ThreadPoolExecutor(max_workers=worker_count) as ex:
        futures = {ex.submit(_process_one_file, path, name_to_id, id_to_name, alt_to_name): path for path in files}
        done = 0
        for fut in as_completed(futures):
            try:
                path, names = fut.result()
            except ImportError as exc:
                missing = getattr(exc, "name", str(exc))
                logging.error("依赖缺失: %s", missing)
                return 2
            except Exception as exc:
                logging.warning("任务执行失败: %s", exc)
                continue

            done += 1
            if names:
                matched_files.append(path)
                for name in sorted(names):
                    rows.append({
                        "命中的内容": name,  # IDs resolved to names
                        "文件名": path.name,
                        "保存路径": str(path),
                    })
            if done % 10 == 0 or done == total:
                logging.info("进度: %d/%d", done, total)

    # 写出 Excel
    try:
        df = pd.DataFrame(rows, columns=["命中的内容", "文件名", "保存路径"]) if rows else pd.DataFrame(
            columns=["命中的内容", "文件名", "保存路径"]
        )
        if not df.empty:
            # Build base filename (without suffix) for deduplication
            df = df.assign(**{"文件名_无后缀": df["文件名"].map(lambda s: Path(str(s)).stem)})
            before = len(df)
            # Deduplicate by (姓名, 文件名去后缀). Ignore different paths.
            df = df.drop_duplicates(subset=["命中的内容", "文件名_无后缀"]).reset_index(drop=True)
            after = len(df)
            logging.info("去重完成: %d -> %d (规则: 姓名+文件名(无后缀))", before, after)
            # Sort by name
            df = df.sort_values(by=["命中的内容"]).reset_index(drop=True)
            logging.info("报表排序: 已按姓名升序排序")
            # Drop helper column before writing
            df = df.drop(columns=["文件名_无后缀"])  # keep original three columns only
        with pd.ExcelWriter(str(report_path), engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="命中明细")
        # Auto-fit columns after writing
        _autosize_excel_columns(report_path, sheet_name="命中明细")
        logging.info("写出报表: %s (命中 %d 条, 文件 %d 个)", report_path, len(df), len(matched_files))
    except Exception as exc:
        logging.error("写出 Excel 失败: %s", exc)
        return 3

    # 复制命中文件
    for src in matched_files:
        try:
            dst = out_dir / src.name
            # 若重名，追加序号避免覆盖
            if dst.exists():
                stem, suf = src.stem, src.suffix
                k = 2
                while True:
                    cand = out_dir / f"{stem}_{k}{suf}"
                    if not cand.exists():
                        dst = cand
                        break
                    k += 1
            shutil.copy2(src, dst)
        except Exception as exc:
            logging.warning("复制失败 %s -> %s: %s", src, out_dir, exc)

    logging.info("完成：共扫描 %d 个文件，命中 %d 个文件，详情见 %s", total, len(matched_files), report_path)
    print(f"完成：共扫描 {total} 个文件，命中 {len(matched_files)} 个文件，输出：{report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
