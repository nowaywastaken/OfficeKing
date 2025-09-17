from __future__ import annotations

"""Utilities for assembling pandas DataFrames and writing Excel reports."""

from typing import Iterable, Tuple

import pandas as pd

from .schema import (
    CLASS_TAG_LABEL,
    CLASS_TAG_SHEET_NAME,
    COLUMN_ACTIVITY_FILE_COUNT,
    COLUMN_ACTIVITY_NAME,
    COLUMN_FILE_PATH,
    COLUMN_MATCH_COUNT,
    COLUMN_MATCH_TOTAL,
    COLUMN_MATCH_TYPE,
    COLUMN_MATCH_VALUE,
    COLUMN_PERSON_ACTIVITY_COUNT,
    COLUMN_PERSON_ACTIVITY_LIST,
    COLUMN_SNIPPET,
    COLUMN_STATUS,
    COLUMN_STUDENT_ID,
    COLUMN_STUDENT_NAME,
    DETAIL_SHEET_NAME,
    PER_ACTIVITY_SHEET_NAME,
    PER_PERSON_SHEET_NAME,
    STUDENT_MATCH_TYPES,
)
from .roster_store import StudentDirectory


def build_report_tables(all_rows: Iterable[dict], roster: StudentDirectory) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Transform raw scan results into the DataFrames needed for reporting."""

    detail_df = pd.DataFrame(list(all_rows))

    student_hits = detail_df[
        (detail_df[COLUMN_STATUS] == "OK")
        & (detail_df[COLUMN_MATCH_COUNT] > 0)
        & (detail_df[COLUMN_MATCH_TYPE].isin(STUDENT_MATCH_TYPES))
    ].copy()

    if not student_hits.empty:
        missing_names = student_hits[COLUMN_STUDENT_NAME].eq("") & student_hits[COLUMN_STUDENT_ID].ne("")
        if missing_names.any():
            student_hits.loc[missing_names, COLUMN_STUDENT_NAME] = student_hits.loc[missing_names, COLUMN_STUDENT_ID].map(
                lambda sid: roster.resolve_name(sid or "") or ""
            )

        grouped = student_hits.groupby([COLUMN_STUDENT_ID, COLUMN_STUDENT_NAME, COLUMN_ACTIVITY_NAME], as_index=False).agg(
            {COLUMN_FILE_PATH: "nunique", COLUMN_MATCH_COUNT: "sum"}
        )
        per_activity = grouped.rename(columns={
            COLUMN_FILE_PATH: COLUMN_ACTIVITY_FILE_COUNT,
            COLUMN_MATCH_COUNT: COLUMN_MATCH_TOTAL,
        })

        per_person = per_activity.groupby([COLUMN_STUDENT_ID, COLUMN_STUDENT_NAME], as_index=False).agg(
            {COLUMN_ACTIVITY_NAME: ["nunique", lambda values: "、".join(sorted(set(values)))]}
        )
        per_person.columns = [
            COLUMN_STUDENT_ID,
            COLUMN_STUDENT_NAME,
            COLUMN_PERSON_ACTIVITY_COUNT,
            COLUMN_PERSON_ACTIVITY_LIST,
        ]
    else:
        per_activity = pd.DataFrame(
            columns=[
                COLUMN_STUDENT_ID,
                COLUMN_STUDENT_NAME,
                COLUMN_ACTIVITY_NAME,
                COLUMN_ACTIVITY_FILE_COUNT,
                COLUMN_MATCH_TOTAL,
            ]
        )
        per_person = pd.DataFrame(
            columns=[
                COLUMN_STUDENT_ID,
                COLUMN_STUDENT_NAME,
                COLUMN_PERSON_ACTIVITY_COUNT,
                COLUMN_PERSON_ACTIVITY_LIST,
            ]
        )

    class_hits = detail_df[
        (detail_df[COLUMN_STATUS] == "OK") & (detail_df[COLUMN_MATCH_TYPE] == CLASS_TAG_LABEL)
    ].copy()

    return detail_df, per_activity, per_person, class_hits


def write_report_workbook(
    output_path: str,
    detail: pd.DataFrame,
    per_activity: pd.DataFrame,
    per_person: pd.DataFrame,
    class_hits: pd.DataFrame,
) -> None:
    """Write all report sheets to an Excel workbook."""

    # Reorder detail columns: name → activity → path, and push hidden columns to the end
    desired_lead = [COLUMN_STUDENT_NAME, COLUMN_ACTIVITY_NAME, COLUMN_FILE_PATH]
    existing = list(detail.columns)
    lead = [c for c in desired_lead if c in existing]
    # Columns we want hidden in the Excel sheet
    hidden_set = {
        COLUMN_STATUS,
        COLUMN_SNIPPET,
        COLUMN_MATCH_TYPE,
        COLUMN_MATCH_VALUE,
        COLUMN_STUDENT_ID,
        COLUMN_MATCH_COUNT,
    }
    rest = [c for c in existing if c not in lead and c not in hidden_set]
    tail_hidden = [
        c
        for c in [
            COLUMN_STATUS,
            COLUMN_SNIPPET,
            COLUMN_MATCH_TYPE,
            COLUMN_MATCH_VALUE,
            COLUMN_STUDENT_ID,
            COLUMN_MATCH_COUNT,
        ]
        if c in existing
    ]
    ordered_cols = lead + rest + tail_hidden if existing else existing
    if ordered_cols:
        try:
            detail = detail.reindex(columns=ordered_cols)
        except Exception:
            pass

    # Sort the detail sheet by student name (then activity and path for stability)
    try:
        if COLUMN_STUDENT_NAME in detail.columns:
            detail[COLUMN_STUDENT_NAME] = detail[COLUMN_STUDENT_NAME].astype(str)
            # Stable merge sort to keep relative order for equal keys
            by_cols = [c for c in [COLUMN_STUDENT_NAME, COLUMN_ACTIVITY_NAME, COLUMN_FILE_PATH] if c in detail.columns]
            if by_cols:
                detail = detail.sort_values(by=by_cols, kind="mergesort", na_position="last")
    except Exception:
        pass

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        detail.to_excel(writer, index=False, sheet_name=DETAIL_SHEET_NAME)
        per_activity.to_excel(writer, index=False, sheet_name=PER_ACTIVITY_SHEET_NAME)
        per_person.to_excel(writer, index=False, sheet_name=PER_PERSON_SHEET_NAME)
        class_hits.to_excel(writer, index=False, sheet_name=CLASS_TAG_SHEET_NAME)

        # Hide selected columns in the detail sheet (match type/value/id/count, status, snippet)
        try:
            from openpyxl.utils import get_column_letter  # type: ignore

            ws = writer.sheets.get(DETAIL_SHEET_NAME)
            if ws is not None and ordered_cols:
                for col_name in (
                    COLUMN_MATCH_TYPE,
                    COLUMN_MATCH_VALUE,
                    COLUMN_STUDENT_ID,
                    COLUMN_MATCH_COUNT,
                    COLUMN_STATUS,
                    COLUMN_SNIPPET,
                ):
                    if col_name in ordered_cols:
                        idx = ordered_cols.index(col_name) + 1  # 1-based for Excel
                        letter = get_column_letter(idx)
                        ws.column_dimensions[letter].hidden = True
        except Exception:
            # If hiding fails (e.g., alternative engine), leave columns visible
            pass

        # Auto-fit column widths for all sheets based on content
        try:
            from openpyxl.utils import get_column_letter  # type: ignore
            import unicodedata as _ud  # type: ignore

            def _disp_len(val: object) -> int:
                if val is None:
                    return 0
                s = str(val)
                best = 0
                for line in s.splitlines():
                    ln = 0
                    for ch in line:
                        ln += 2 if _ud.east_asian_width(ch) in ("W", "F") else 1
                    if ln > best:
                        best = ln
                return best

            def _autofit(ws) -> None:
                # Look at up to 1000 data rows for performance
                max_rows = min(ws.max_row or 0, 1000)
                for col_idx in range(1, (ws.max_column or 0) + 1):
                    letter = get_column_letter(col_idx)
                    # Skip hidden columns (keep them compact)
                    cd = ws.column_dimensions.get(letter)
                    if getattr(cd, "hidden", False):
                        continue
                    max_len = 0
                    # Header cell
                    hdr = ws.cell(row=1, column=col_idx).value
                    max_len = max(max_len, _disp_len(hdr))
                    # Data cells
                    for row_idx in range(2, max_rows + 1):
                        v = ws.cell(row=row_idx, column=col_idx).value
                        if v is not None:
                            max_len = max(max_len, _disp_len(v))
                    # Convert to Excel width units with padding, and clamp
                    width = min(80.0, max(8.0, max_len * 1.2 + 2.0))
                    ws.column_dimensions[letter].width = width

            for name, ws in writer.sheets.items():
                _autofit(ws)
        except Exception:
            # Auto-fit is best-effort; ignore if engine/sheet not available
            pass
